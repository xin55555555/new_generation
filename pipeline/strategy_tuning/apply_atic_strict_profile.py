#!/usr/bin/env python3
"""Apply the ATIC strict strategy profile to incremental templates."""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline_settings import DEFAULT_TEMPLATE_DIR, PROJECT_ROOT


DEFAULT_BASELINE = PROJECT_ROOT / "policy_t.json"
DEFAULT_ATIC_XLSX = PROJECT_ROOT / "ATIC策略配置基线20260602.xlsx"
DEFAULT_SHEET = "ATIC策略配置-R25.0"
STRICT_COLUMN_NAME = "严格-Strict"


def spec(**kwargs: Any) -> dict[str, Any]:
    return kwargs


# The source row numbers point to sheet ATIC策略配置-R25.0, column 严格-Strict.
# Rule thresholds are retained for audit/reporting only; generated templates always
# inherit threshold values from the baseline policy_t.json.
STRICT_PROFILE: dict[str, list[dict[str, Any]]] = {
    "SYN_FLOOD": [
        spec(attack_type="Tcp_Syn_Flood", enable_status=1, threshold="2000", source_rows=[5, 7]),
        spec(attack_type="Tcp_Syn_Flood_Bandwidth", enable_status=1, threshold="10000", source_rows=[8]),
        spec(attack_type="Tcp_Syn_First_Packet_Check", enable_status=1, threshold="0.5-4", source_rows=[10, 11]),
        spec(attack_type="Syn_Flood_Src_Challenge_Auth", enable_status=1, source_rows=[12, 13]),
        spec(attack_type="Syn_Advanced_Detection", enable_status=1, source_rows=[15]),
        spec(attack_type="Tcp_Syn_Abn_Src_Limit", enable_status=1, threshold="10,1", source_rows=[17]),
        spec(attack_type="Syn_Flood_Abnormal_Src_Block", enable_status=1, threshold="80,10,1,2,3", source_rows=[18, 19, 20, 21]),
    ],
    "TCP_SYNACK_FLOOD": [
        spec(attack_type="Tcp_SynAck_Flood", enable_status=1, threshold="2000", source_rows=[22, 24]),
        spec(attack_type="Tcp_SynAck_First_Packet_Check", enable_status=1, threshold="1-4", source_rows=[26, 27]),
    ],
    "ACK_FLOOD": [
        spec(attack_type="Tcp_Ack_Flood_Attack", enable_status=1, threshold="20000", source_rows=[30, 32]),
        spec(attack_type="Tcp_Ack_Flood_Attack_Bandwidth", enable_status=1, threshold="80000", source_rows=[33]),
        spec(attack_type="Tcp_Ack_First_Packet_Check", enable_status=1, threshold="0-6", source_rows=[35, 36]),
        spec(attack_type="Tcp_Ack_Session_Detection_Mode", enable_status=1, mode_type=0, source_rows=[37, 38]),
        spec(attack_type="Tcp_Ack_Src_Low_Speed_Behavior_Analysis", enable_status=1, threshold="3", source_rows=[42, 48]),
        spec(attack_type="Tcp_Ack_Session_Behavior_Analysis", enable_status=1, threshold="300", source_rows=[49, 50]),
        spec(attack_type="Tcp_Abn_Ack_Conn_Detect", enable_status=1, threshold="1000,90,15", source_rows=[51, 52, 53, 54]),
    ],
    "TCP_FLOOD": [
        spec(attack_type="Tcp_Abnormal", enable_status=1, threshold="500", source_rows=[3, 4]),
        spec(attack_type="Tcp_Fin_Flood", enable_status=1, threshold="2000", source_rows=[63, 65]),
        spec(attack_type="Tcp_Fragment_Traffic_Limiting", enable_status=1, threshold="10000", source_rows=[60, 93]),
        spec(attack_type="Tcp_New_Conn_Limiting", enable_status=1, threshold="100000", source_rows=[95, 96]),
    ],
    "TCP_STATE_EXHAUSTION": [
        spec(attack_type="Tcp_Concurrent_Connection_Check_By_Destination_Ip", enable_status=1, threshold="20000", source_rows=[66, 68]),
        spec(attack_type="Tcp_New_Connection_Rate_Check_By_Destination_Ip", enable_status=1, threshold="5000", source_rows=[69]),
        spec(attack_type="Tcp_Connection_Number_Check_For_Source_Ip", enable_status=1, threshold="100", source_rows=[71, 72]),
        spec(attack_type="Tcp_Src_Null_Conn_Detect", enable_status=1, threshold="5,5", source_rows=[76, 77]),
        spec(attack_type="Tcp_Src_Abn_Conn_Drop", enable_status=1, threshold="2,5", source_rows=[83]),
    ],
    "UDP_FLOOD": [
        spec(attack_type="Udp_Abnormal", enable_status=1, threshold="500", source_rows=[101, 103]),
        spec(attack_type="Udp_Traffic_Limiting", enable_status=1, threshold="50000", source_rows=[123, 124]),
        spec(attack_type="Udp_Traffic_Limiting_Strict", enable_status=1, source_rows=[125]),
        spec(attack_type="Udp_Fragment_Rate_Limiting", enable_status=1, threshold="10000", source_rows=[127]),
        spec(attack_type="Udp_New_Conn_Limiting", enable_status=1, threshold="100000", source_rows=[128, 130]),
        spec(attack_type="Udp_Session_Behavior_Detect_Packet_Interval", enable_status=1, threshold="0.1-2", source_rows=[108, 109]),
        spec(attack_type="Udp_Correlation", enable_status=1, source_rows=[110, 111, 114]),
    ],
    "UDP_REFLECTION": [
        spec(attack_type="Udp_Traffic_Limiting", enable_status=1, threshold="50000", source_rows=[123, 124]),
        spec(attack_type="Udp_Fragment_Rate_Limiting", enable_status=1, threshold="10000", source_rows=[127]),
        spec(attack_type="Udp_New_Conn_Limiting", enable_status=1, threshold="100000", source_rows=[128, 130]),
        spec(attack_type="Statistics_Based_On_Destination_Ip", enable_status=1, threshold="1000", source_rows=[321, 323, 325]),
    ],
    "UDP_AMPLIFICATION": [
        spec(attack_type="Udp_Traffic_Limiting", enable_status=1, threshold="50000", source_rows=[123, 124]),
        spec(attack_type="Udp_Fragment_Rate_Limiting", enable_status=1, threshold="10000", source_rows=[127]),
        spec(attack_type="Udp_New_Conn_Limiting", enable_status=1, threshold="100000", source_rows=[128, 130]),
        spec(attack_type="Statistics_Based_On_Destination_Ip", enable_status=1, threshold="1000", source_rows=[321, 323, 325]),
    ],
    "AMPLIFICATION_REQUEST": [],
    "ICMP_FLOOD": [
        spec(attack_type="Icmp_Traffic_Limiting", enable_status=1, threshold="500", source_rows=[291, 292]),
    ],
    "GRE_FLOOD": [
        spec(attack_type="Other_Traffic_Limiting", enable_status=1, threshold="5000", source_rows=[304, 306, 308]),
        spec(attack_type="Other_New_Conn_Limiting", enable_status=1, threshold="10000", source_rows=[310, 312]),
    ],
    "IP_FLOOD": [
        spec(attack_type="Other_Traffic_Limiting", enable_status=1, threshold="5000", source_rows=[304, 306, 308]),
        spec(attack_type="Other_New_Conn_Limiting", enable_status=1, threshold="10000", source_rows=[310, 312]),
    ],
    "HTTP_APP": [
        spec(attack_type="Http_Abn_Src_Block", enable_status=1, threshold="60,10,5", source_rows=[162, 164, 165, 166]),
        spec(attack_type="Http_Single_Uri_Detect", enable_status=1, threshold="60,10,5", source_rows=[162, 164, 165, 166]),
        spec(attack_type="Http_Large_Resource_Detect", enable_status=1, threshold="100,60,10,5", source_rows=[179, 180, 182, 183, 184]),
        spec(attack_type="Http_Slowloris_Detect", enable_status=1, threshold="1000,50,10", source_rows=[201, 202, 204]),
        spec(attack_type="Http_Abn_Conn_Defense", enable_status=1, threshold="10000", source_rows=[196, 198]),
        spec(attack_type="Http_Null_Conn_Detect", enable_status=1, threshold="1,5", source_rows=[206, 207]),
        spec(attack_type="Http_Range_Amp_Detect", enable_status=1, threshold="2,5", source_rows=[208, 209]),
        spec(attack_type="Http_Multi_Method_Conn_Detect", enable_status=1, threshold="2,5", source_rows=[210, 211]),
    ],
    "HTTPS_APP": [
        spec(attack_type="Tls_Attack_Dst_Ip", enable_status=1, threshold="20000", source_rows=[215, 217]),
        spec(attack_type="Tls_Src_Challenge_Auth", enable_status=1, source_rows=[220]),
        spec(attack_type="Tls_Unfixed_Resource_Detect", enable_status=1, threshold="200,1,1", source_rows=[221, 222, 223]),
        spec(attack_type="Tls_FingerPrint_Detect", enable_status=1, threshold="5,60,2000,3", source_rows=[224, 226, 227, 229]),
        spec(attack_type="Tls_Large_Resource_Detect", enable_status=1, threshold="100,60,10,5", source_rows=[236, 238, 240, 241, 242]),
        spec(attack_type="Tls_Fixed_Resource_Detect", enable_status=1, threshold="60,10,5", source_rows=[243, 244, 245, 246]),
        spec(attack_type="Tls_Session_Attack_Defense", enable_status=1, threshold="10000", source_rows=[249, 251]),
        spec(attack_type="Tls_Null_Conn_Detect", enable_status=1, threshold="1,5", source_rows=[254, 255]),
        spec(attack_type="Tls_Incomplete_Conn_Detect", enable_status=1, threshold="2,5", source_rows=[256, 257]),
        spec(attack_type="Tls_Rst_Conn_Detect", enable_status=1, threshold="2,5", source_rows=[260, 261]),
    ],
    "QUIC_APP": [
        spec(attack_type="Udp_Traffic_Limiting", enable_status=1, threshold="50000", source_rows=[123, 124]),
        spec(attack_type="Udp_New_Conn_Limiting", enable_status=1, threshold="100000", source_rows=[128, 130]),
    ],
    "DNS_REQUEST_FLOOD": [
        spec(attack_type="Dns_Abnormal", enable_status=1, threshold="500", source_rows=[263, 265]),
        spec(attack_type="Dns_Query_Flood", enable_status=1, threshold="2000", mode_type=2, source_rows=[266, 268, 269, 270]),
        spec(attack_type="Dns_Query_First_Packet_Check", enable_status=1, threshold="1-5", source_rows=[272, 273]),
        spec(attack_type="Dns_Src_Query_Limiting", enable_status=1, threshold="100", source_rows=[286, 287]),
    ],
    "DNS_REPLY_FLOOD": [],
    "DNS_APP": [
        spec(attack_type="Dns_Abnormal", enable_status=1, threshold="500", source_rows=[263, 265]),
        spec(attack_type="Dns_Query_Flood", enable_status=1, threshold="2000", mode_type=2, source_rows=[266, 268, 269, 270]),
        spec(attack_type="Dns_Query_First_Packet_Check", enable_status=1, threshold="1-5", source_rows=[272, 273]),
        spec(attack_type="Dns_Src_Query_Limiting", enable_status=1, threshold="100", source_rows=[286, 287]),
    ],
    "SIP_APP": [],
    "DTLS_APP": [
        spec(attack_type="Udp_Traffic_Limiting", enable_status=1, threshold="50000", source_rows=[123, 124]),
        spec(attack_type="Udp_New_Conn_Limiting", enable_status=1, threshold="100000", source_rows=[128, 130]),
    ],
    "UNKNOWN_DDOS": [
        spec(attack_type="Other_Traffic_Limiting", enable_status=1, threshold="5000", source_rows=[304, 306, 308]),
        spec(attack_type="Other_New_Conn_Limiting", enable_status=1, threshold="10000", source_rows=[310, 312]),
    ],
}


UNREPRESENTED_NOTES = [
    {
        "topic": "SYN source challenge mode",
        "source_rows": [13],
        "strict_value": "智能模式",
        "reason": "Current JSON schema only exposes mode_type; keep baseline mode_type unless an explicit device enum is confirmed.",
    },
    {
        "topic": "UDP fingerprint/watermark details",
        "source_rows": [110, 111, 113, 114, 118],
        "strict_value": "指纹学习/一致性学习开启，水印关闭",
        "reason": "Current JSON has Udp_Correlation but no direct one-to-one fields for all fingerprint-learning parameters.",
    },
    {
        "topic": "DNS reply flood",
        "source_rows": [277],
        "strict_value": "关闭",
        "reason": "Strict profile intentionally produces an empty DNS_REPLY_FLOOD template unless a deployment-specific override is required.",
    },
    {
        "topic": "SIP flood",
        "source_rows": [295, 300],
        "strict_value": "关闭",
        "reason": "Strict profile intentionally produces an empty SIP_APP template.",
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--baseline", default=str(DEFAULT_BASELINE), help="Default policy JSON")
    parser.add_argument("--xlsx", default=str(DEFAULT_ATIC_XLSX), help="ATIC strategy workbook")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Workbook sheet containing strict strategy")
    parser.add_argument("--template-dir", default=str(DEFAULT_TEMPLATE_DIR), help="Incremental template directory")
    parser.add_argument("--report", default=None, help="Optional output report path")
    parser.add_argument("--backup-dir", default=None, help="Optional backup directory before rewriting templates")
    return parser.parse_args()


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def baseline_policy_index(baseline: dict[str, Any]) -> dict[str, dict[str, dict[str, Any]]]:
    index: dict[str, dict[str, dict[str, Any]]] = {}
    for device in baseline.get("policyDetailDto", []):
        device_ip = device.get("deviceIp")
        if not isinstance(device_ip, str):
            continue
        index[device_ip] = {
            policy["attack_type"]: policy
            for policy in device.get("policy", [])
            if isinstance(policy, dict) and isinstance(policy.get("attack_type"), str)
        }
    return index


def strict_row_values(xlsx: Path, sheet_name: str) -> dict[int, Any]:
    workbook = load_workbook(xlsx, data_only=True)
    worksheet = workbook[sheet_name]
    header = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    try:
        strict_col = header.index(STRICT_COLUMN_NAME) + 1
    except ValueError as exc:
        raise RuntimeError(f"sheet {sheet_name} does not contain {STRICT_COLUMN_NAME}") from exc
    return {row: worksheet.cell(row, strict_col).value for row in range(2, worksheet.max_row + 1)}


def build_policy_item(baseline_policy: dict[str, Any], rule: dict[str, Any]) -> dict[str, Any]:
    item = copy.deepcopy(baseline_policy)
    # Thresholds must remain the device-default values from the baseline template.
    # The strict profile only decides which switches/modes should be enabled.
    for key in ("enable_status", "mode_type"):
        if key in rule:
            item[key] = rule[key]
    return item


def policy_counts(payload: dict[str, Any]) -> dict[str, int]:
    return {
        str(device.get("deviceIp")): len(device.get("policy", []))
        for device in payload.get("policyDetailDto", [])
        if isinstance(device, dict)
    }


def update_template(
    path: Path,
    baseline_path: Path,
    xlsx_path: Path,
    sheet_name: str,
    baseline_index: dict[str, dict[str, dict[str, Any]]],
    strict_values: dict[int, Any],
    backup_dir: Path,
) -> dict[str, Any]:
    payload = load_json(path)
    before_counts = policy_counts(payload)
    attack_type = payload.get("normalized_attack_type")
    rules = STRICT_PROFILE.get(str(attack_type), [])

    new_devices = []
    missing_policies = []
    applied_rules = []
    removed_equal = 0
    for device_ip, policies in baseline_index.items():
        new_policies = []
        for rule in rules:
            policy_name = rule["attack_type"]
            baseline_policy = policies.get(policy_name)
            if baseline_policy is None:
                missing_policies.append(policy_name)
                continue
            item = build_policy_item(baseline_policy, rule)
            applied_rules.append({
                "attack_type": policy_name,
                "source_rows": rule.get("source_rows", []),
                "strict_values": {
                    str(row): strict_values.get(row)
                    for row in rule.get("source_rows", [])
                },
                "enable_status": item.get("enable_status"),
                "threshold": item.get("threshold"),
                "strict_profile_threshold_ignored": rule.get("threshold"),
                "mode_type": item.get("mode_type"),
            })
            if item == baseline_policy:
                removed_equal += 1
                continue
            new_policies.append(item)
        if new_policies:
            new_devices.append({"deviceIp": device_ip, "policy": new_policies})

    backup_path = backup_dir / path.name
    shutil.copy2(path, backup_path)

    payload.pop("source_policy_override", None)
    payload["comparison_baseline"] = str(baseline_path.resolve())
    payload["strict_profile_source"] = {
        "xlsx": str(xlsx_path.resolve()),
        "sheet": sheet_name,
        "column": STRICT_COLUMN_NAME,
    }
    payload["policyDetailDto"] = new_devices
    write_json(path, payload)

    after_counts = policy_counts(payload)
    return {
        "template": str(path.resolve()),
        "backup": str(backup_path.resolve()),
        "normalized_attack_type": attack_type,
        "before_policy_count": sum(before_counts.values()),
        "after_policy_count": sum(after_counts.values()),
        "before_devices": before_counts,
        "after_devices": after_counts,
        "rules_requested": len(rules),
        "rules_applied": len(applied_rules),
        "removed_equal_to_baseline": removed_equal,
        "missing_baseline_policies": sorted(set(missing_policies)),
        "applied_rules": applied_rules,
    }


def update_manifest(template_dir: Path, baseline_path: Path, xlsx_path: Path, sheet_name: str) -> None:
    manifest_path = template_dir / "switch_template_manifest.json"
    if not manifest_path.exists():
        return
    manifest = load_json(manifest_path)
    meta = manifest.setdefault("meta", {})
    meta["base_template"] = str(baseline_path.resolve())
    meta["strict_profile_source"] = {
        "xlsx": str(xlsx_path.resolve()),
        "sheet": sheet_name,
        "column": STRICT_COLUMN_NAME,
    }
    strategy = meta.setdefault("strategy", {})
    strategy["template_mode"] = "incremental_patch"
    strategy["comparison_baseline"] = str(baseline_path.resolve())
    strategy["threshold_policy"] = "threshold values are always inherited from the baseline template; ATIC strict profile only changes switches/modes"

    by_attack_type = {}
    for path in template_dir.glob("*.switch_template.incremental.json"):
        payload = load_json(path)
        by_attack_type[payload.get("normalized_attack_type")] = payload

    for item in manifest.get("templates", []):
        item.pop("source_policy_override", None)
        notes = item.get("notes")
        if isinstance(notes, str):
            notes = notes.replace(" Override from policy_ack_flood.json against policy_init baseline.", "")
            notes = notes.replace(" Override from policy_udp_flood.json against policy_init baseline.", "")
            item["notes"] = notes
        payload = by_attack_type.get(item.get("normalized_attack_type"))
        if payload is None:
            continue
        changes = []
        null_threshold = []
        for device in payload.get("policyDetailDto", []):
            for policy in device.get("policy", []):
                changes.append({
                    "deviceIp": device.get("deviceIp"),
                    "attack_type": policy.get("attack_type"),
                    "new_enable_status": policy.get("enable_status"),
                    "threshold": policy.get("threshold"),
                    "mode_type": policy.get("mode_type"),
                })
                if policy.get("enable_status") == 1 and policy.get("threshold") is None:
                    null_threshold.append(policy.get("attack_type"))
        item["changed_entry_count"] = len(changes)
        item["changes"] = changes
        item["enabled_with_null_threshold"] = sorted(set(x for x in null_threshold if x))
    write_json(manifest_path, manifest)


def main() -> int:
    args = parse_args()
    baseline_path = Path(args.baseline).expanduser().resolve()
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    template_dir = Path(args.template_dir).expanduser().resolve()
    if not baseline_path.is_file():
        raise SystemExit(f"baseline not found: {baseline_path}")
    if not xlsx_path.is_file():
        raise SystemExit(f"xlsx not found: {xlsx_path}")
    if not template_dir.is_dir():
        raise SystemExit(f"template dir not found: {template_dir}")

    timestamp = dt.datetime.now(dt.UTC).strftime("%Y%m%dT%H%M%SZ")
    backup_dir = (
        Path(args.backup_dir).expanduser().resolve()
        if args.backup_dir
        else template_dir / f"backup_before_atic_strict_{timestamp}"
    )
    backup_dir.mkdir(parents=True, exist_ok=True)

    baseline = load_json(baseline_path)
    baseline_index = baseline_policy_index(baseline)
    strict_values = strict_row_values(xlsx_path, args.sheet)
    reports = [
        update_template(
            path=path,
            baseline_path=baseline_path,
            xlsx_path=xlsx_path,
            sheet_name=args.sheet,
            baseline_index=baseline_index,
            strict_values=strict_values,
            backup_dir=backup_dir,
        )
        for path in sorted(template_dir.glob("*.switch_template.incremental.json"))
    ]
    update_manifest(template_dir, baseline_path, xlsx_path, args.sheet)

    report = {
        "baseline": str(baseline_path),
        "strict_profile_source": {
            "xlsx": str(xlsx_path),
            "sheet": args.sheet,
            "column": STRICT_COLUMN_NAME,
        },
        "template_dir": str(template_dir),
        "backup_dir": str(backup_dir),
        "template_count": len(reports),
        "total_before_policy_count": sum(item["before_policy_count"] for item in reports),
        "total_after_policy_count": sum(item["after_policy_count"] for item in reports),
        "unrepresented_notes": UNREPRESENTED_NOTES,
        "templates": reports,
    }
    report_path = (
        Path(args.report).expanduser().resolve()
        if args.report
        else template_dir / "atic_strict_profile_report.json"
    )
    write_json(report_path, report)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
